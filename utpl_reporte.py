import time
import openpyxl
import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import ElementNotVisibleException
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.support.ui import Select
#
import pytest
import time
import json
#Para reportes
import unittest
import HTMLTestRunner

#####
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
#Para encontrar elementos que no son visibles
from selenium.webdriver.common.action_chains import ActionChains

#Permite recuperar las filas del excel
def recu_filas_excel(hoja, inicio_fila = 0):
	list_estu = []
	index_fila = 1
	for filas in range(hoja.max_row):
		if index_fila > inicio_fila and hoja.cell(row=index_fila, column=1).value != None:
			regi_estu = {
				'nume_fila':str(index_fila),
				'TipoIden': str(hoja.cell(row=index_fila, column=1).value),
				'identifica': str(hoja.cell(row=index_fila, column=2).value),
				'nombre': str(hoja.cell(row=index_fila, column=3).value),
				'apellido1': str(hoja.cell(row=index_fila, column=4).value),
				'apellido2': str(hoja.cell(row=index_fila, column=5).value),
				'nacimiento': str(hoja.cell(row=index_fila, column=6).value),
				'esta_civi_estu': str(hoja.cell(row=index_fila, column=7).value),
				'etni_estu': str(hoja.cell(row=index_fila, column=8).value),
				'sexo': str(hoja.cell(row=index_fila, column=9).value),
				'pais': str(hoja.cell(row=index_fila, column=10).value),
				'ciudad': str(hoja.cell(row=index_fila, column=11).value),
				'calle1': str(hoja.cell(row=index_fila, column=12).value),
				'calle2': str(hoja.cell(row=index_fila, column=13).value),
				'ncasa': str(hoja.cell(row=index_fila, column=14).value),
				'referencia': str(hoja.cell(row=index_fila, column=15).value),
				'celular': str(hoja.cell(row=index_fila, column=16).value),
				'fijo': str(hoja.cell(row=index_fila, column=17).value),
				'correo': str(hoja.cell(row=index_fila, column=18).value),
				'contrasenia': str(hoja.cell(row=index_fila, column=19).value),
				'centro': str(hoja.cell(row=index_fila, column=20).value),
				'carrera': str(hoja.cell(row=index_fila, column=21).value),

			}
			#regi_estu = depurar_textos(regi_estu)

			list_estu.append(regi_estu)
		else:
			if hoja.cell(row=index_fila, column=1).value == None:
				break
		index_fila+=1

	return list_estu


#Completa los campos sin texto, con valores por defecto
def depurar_textos(estu):
	if (estu['nume_serv'] == '' or estu['nume_serv'] == 'None'):
		estu['nume_serv'] = 'S/N'

	if (estu['deta_serv'] == '' or estu['deta_serv'] == 'None'):
		estu['deta_serv'] = 'S/N'

	if (estu['esta_serv'] == '' or estu['esta_serv'] == 'None'):
		estu['esta_serv'] = 'S/N'
	else:
		estu['esta_serv'] = estu['esta_serv'].upper()

	return estu


#Obtiene la fecha actual
def obte_fech_actu(formato = 'REPORTE'):
	fech_actu = ''
	fech_carg = datetime.datetime.now()
	year = fech_carg.year
	month = fech_carg.month
	day = fech_carg.day
	hour = fech_carg.hour
	minute = fech_carg.minute

	if formato == 'REPORTE':
		fech_actu = str(year)+"_"+str(month)+"_"+str(day)+"_"+str(hour)+"_"+str(minute)

	return fech_actu


#Permite buscar una forma por su codigo y acceder a ella
def buscar_acceder_forma(driver, nomb_forma, code, estu):
	driver.switch_to.default_content()
	lupa = driver.find_element_by_id('sidebarSearchLink')
	lupa.click()
	time.sleep(1)
	busc_prin = driver.find_element_by_id('search')
	busc_prin.send_keys(nomb_forma)
	busc_prin.send_keys(u'\ue007') #Preciona ENTER
	if revisar_elem("//iframe[@id='bannerHS']", code, estu) == False:
		return False
	else:
		driver.switch_to.frame(driver.find_element_by_xpath("//iframe[@id='bannerHS']"))
		return True




#Permite ingresar al sistema
def login_site(driver):
	driver.get(url_ambiente)
	time.sleep(3)
	user_box = driver.find_element_by_xpath("//input[@name='UserName']")
	pass_box = driver.find_element_by_xpath("//input[@name='Password']")
	subm_butt = driver.find_element_by_id('submitButton')
	time.sleep(3)



#Escribe los errores dentro del registro de log
def escribir_error_log(estu, code, mensaje_error):
	f.write(
			str(estu['nume_fila'])+","+
			str(estu['identifica'])+","+
			mensaje_error +" CODIGO: "+ code+"\n"
		)
	print(
		str(estu['nume_fila'])+","+
		str(estu['identifica'])+", "+
		mensaje_error+" CODIGO "+code
	)


#Estibe los exitos dentro del log
def escribir_registro_exitoso_log(estu):
	f.write(
			str(estu['nume_fila'])+","+
			str(estu['id_estu'])+","+
			"REGISTRO EXITOSO \n"
		)
	print(
		str(estu['nume_fila'])+" | ID: "+
		str(estu['id_estu'])+" | "+
		"ESTADO: OK"
	)


#Permite enviar un texto a un input mediante codigo python, con tiempo de espera al inicio
def enviar_teclas_por_id(teclas, id_elemento, espera_previa = 0):
	if revisar_elem_por_id(id_elemento, estu, "INTENTO DE CLICK") == False:
		raise NoSuchElementException
	if espera_previa > 0:
		time.sleep(espera_previa)
	driver.execute_script("document.getElementById('"+id_elemento+"').click()")
	driver.find_element_by_id(id_elemento).send_keys(""+teclas)



#Permite enviar un texto a un input mediante codigo python, con tiempo de espera a la mitad del proceso inicio
def enviar_teclas_por_id_desc_inter(teclas, id_elemento, espera_previa = 0, usar_enter = False):
	if revisar_elem_por_id(id_elemento, estu, "INTENTO DE CLICK") == False:
		raise NoSuchElementException
	driver.execute_script("document.getElementById('"+id_elemento+"').click()")
	if espera_previa > 0:
		time.sleep(espera_previa)
	if usar_enter:
		driver.find_element_by_id(id_elemento).send_keys(""+teclas, Keys.ENTER)
	else:
		driver.find_element_by_id(id_elemento).send_keys(""+teclas)



#Permite enviar un texto a un input mediante codigo javascript, con tiempo de espera al inicio
def enviar_texto_js_por_id(texto, id_elemento, espera_previa = 0):
	if revisar_elem_por_id(id_elemento, estu, "INTENTO DE CLICK") == False:
		raise NoSuchElementException
	if espera_previa > 0:
		time.sleep(espera_previa)
	driver.execute_script("document.getElementById('"+id_elemento+"').click()")
	driver.execute_script("document.getElementById('"+id_elemento+"').value = '"+texto+"'")


#Permite enviar un texto a un input mediante codigo javascript, con tiempo de espera a la mitad del proceso
def enviar_texto_js_por_id_desc_inter(texto, id_elemento, espera_previa = 0):
	if revisar_elem_por_id(id_elemento, estu, "INTENTO DE CLICK") == False:
		raise NoSuchElementException
	driver.execute_script("document.getElementById('"+id_elemento+"').click()")
	if espera_previa > 0:
		time.sleep(espera_previa)
	driver.execute_script("document.getElementById('"+id_elemento+"').value = '"+texto+"'")

# REFERENCIA PARA CAMBIAR DE FRAME (VENTANA INTERNA)
#driver.switch_to.frame(driver.find_element_by_xpath("//iframe[@id='bannerHS']"))

def login_site_test(driver):
	driver.get(url_ambiente)
	time.sleep(3)
	#import pdb; pdb.set_trace()

	#encontrar elemnto en Consola con XPah
	#document.evaluate("//button[@id='closebaner']", document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;  
	##driver.find_element_by_id('closebaner').click()
	#driver.execute_script("document.getElementById('closebaner').click()")
	#driver.execute_script("document.getElementById('closebaner')").click()
	#driver.find_element_by_xpath("//button[@id='closebaner']").click()
	#driver.execute_script("document.evaluate('//button[@id=\"closebaner\"]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.click();")



#Permite un tiempo de espera activo y busqueda de elemento por id
def revisar_elem_por_id(id_elem, estu, code, nume_inte = 20, tiem_espe = 2):
	intentos_restante = nume_inte - 1
	for x in range(1,nume_inte):
		try :
			driver.find_element_by_id(id_elem)
			return True
		except NoSuchElementException as e:
			if (x < intentos_restante ):
				time.sleep(tiem_espe)
				print("Esperando "+str(tiem_espe*x)+" "+id_elem)
			else:
				mensaje_error = "ERROR: Tiempo de espera agotado, "+str(e)
				escribir_error_log(estu, code+" ELEMENTO: "+id_elem, mensaje_error)


#Permite un tiempo de espera activo y busqueda de elemento por css
def revisar_elem_por_css(css_elem, estu, code, nume_inte = 20, tiem_espe = 2):
	intentos_restante = nume_inte - 1
	for x in range(1,nume_inte):
		try :
			driver.find_element_by_css_selector(css_elem)
			return True
		except NoSuchElementException as e:
			if (x < intentos_restante ):
				time.sleep(tiem_espe)
				print("Esperando "+str(tiem_espe*x)+" "+css_elem)
			else:
				mensaje_error = "ERROR: Tiempo de espera agotado, "+str(e)
				escribir_error_log(estu, code+" ELEMENTO: "+css_elem, mensaje_error)
				return False


#Permite un tiempo de espera activo y busqueda de elemento por xpath
def revisar_elem(elem_xpath, code, estu, nume_inte = 30, tiem_espe = 3):
	intentos_restante = nume_inte - 1
	for x in range(1,nume_inte):
		try :
			driver.find_element_by_xpath(elem_xpath)
			return True
		except NoSuchElementException as e:
			if (x < intentos_restante ):
				time.sleep(tiem_espe)
				print("Esperando "+str(tiem_espe*x)+" "+elem_xpath)
			else:
				mensaje_error = "ERROR: Tiempo de espera agotado: "+str(e)
				escribir_error_log(estu, code, mensaje_error)
				return False


#Agregara la identificacion del estudiante en el formulario
def agre_iden(estu):	
	#driver.get("https://srv-si-001.utpl.edu.ec/INSCRIPCION_NUEVOS_MAD/")
	driver.set_window_size(1552, 840)
	driver.find_element(By.ID, "closebaner").click()
	driver.find_element(By.ID, "person-identificationTypeId").click()
	tipo_iden=estu['TipoIden']
	dropdown = driver.find_element(By.ID, "person-identificationTypeId")
	dropdown.find_element(By.XPATH, "//option[. = '"+tipo_iden+"']").click()
	driver.find_element(By.ID, "person-identificationTypeId").click()
	driver.find_element(By.ID, "person-identification").click()
	driver.find_element(By.ID, "person-identification").send_keys(estu['identifica'])
	driver.implicitly_wait(10) # seconds
	driver.find_element(By.CSS_SELECTOR, ".form-group > .btn").click()
	driver.find_element(By.ID, "person-fullNames").click()
	driver.find_element(By.ID, "person-fullNames").send_keys(estu['nombre'])
	driver.find_element(By.ID, "person-lastName").click()
	driver.find_element(By.ID, "person-lastName").send_keys(estu['apellido1'])
	driver.find_element(By.ID, "person-secondSurname").click()
	driver.find_element(By.ID, "person-secondSurname").send_keys(estu['apellido2'])
	driver.find_element(By.NAME, "myDate").click()
	driver.find_element(By.NAME, "myDate").send_keys(estu['nacimiento'])
	estadocivil=estu['esta_civi_estu'] ##igualo una variable al valor de el excel
	driver.find_element(By.ID, "person-civilStatusId").click()
	dropdown = driver.find_element(By.ID, "person-civilStatusId")
	dropdown.find_element(By.XPATH, "//option[. = '"+estadocivil+"']").click()
	driver.find_element(By.ID, "person-civilStatusId").click()
	etnia=estu['etni_estu']
	driver.find_element(By.ID, "person-ethnicityId").click()
	dropdown = driver.find_element(By.ID, "person-ethnicityId")
	dropdown.find_element(By.XPATH, "//option[. = '"+etnia+"']").click()
	driver.find_element(By.ID, "person-ethnicityId").click()
	sexoper=estu['sexo']
	driver.find_element(By.ID, "person-sex").click()
	dropdown = driver.find_element(By.ID, "person-sex")
	dropdown.find_element(By.XPATH, "//option[. = '"+sexoper+"']").click()
	driver.find_element(By.ID, "person-sex").click()
	driver.find_element(By.CSS_SELECTOR, "main").click()
	paispers = estu['pais']
	driver.find_element(By.ID, "person-countryId").click()
	dropdown = driver.find_element(By.ID, "person-countryId")
	dropdown.find_element(By.XPATH, "//option[. = '"+paispers+"']").click()
	driver.find_element(By.ID, "person-countryId").click()
	driver.find_element(By.ID, "person-city").click()
	driver.find_element(By.ID, "person-city").send_keys(estu['ciudad'])
	driver.find_element(By.ID, "person-mainStreet").click()
	driver.find_element(By.ID, "person-mainStreet").send_keys(estu['calle1'])
	driver.find_element(By.ID, "person-secondaryStreet").click()
	driver.find_element(By.ID, "person-secondaryStreet").click()
	driver.find_element(By.ID, "person-secondaryStreet").send_keys(estu['calle2'])
	driver.find_element(By.ID, "person-houseNumber").click()
	driver.find_element(By.ID, "person-houseNumber").send_keys(estu['ncasa'])
	driver.find_element(By.ID, "person-reference").click()
	driver.find_element(By.ID, "person-reference").send_keys(estu['referencia'])
	driver.find_element(By.ID, "person-cellPhone").click()
	driver.find_element(By.ID, "person-cellPhone").send_keys(estu['celular'])
	driver.find_element(By.ID, "person-homePhone").click()
	driver.find_element(By.ID, "person-homePhone").send_keys(estu['fijo'])
	driver.find_element(By.CSS_SELECTOR, ".form-row:nth-child(14)").click()
	driver.find_element(By.ID, "person-userEmail").click()
	driver.find_element(By.ID, "person-userEmail").send_keys(estu['correo'])
	driver.find_element(By.ID, "person-emailConfirm").click()
	driver.find_element(By.ID, "person-emailConfirm").send_keys(estu['correo'])
	driver.find_element(By.ID, "person-password").click()
	driver.find_element(By.ID, "person-password").send_keys(estu['contrasenia'])
	driver.find_element(By.ID, "person-passwordConfirm").click()
	driver.find_element(By.ID, "person-passwordConfirm").send_keys(estu['contrasenia'])
	driver.find_element(By.CSS_SELECTOR, ".form-row:nth-child(1)").click()
	centroun=estu['centro']
	driver.find_element(By.ID, "person-universityCenterId").click()
	dropdown = driver.find_element(By.ID, "person-universityCenterId")
	dropdown.find_element(By.XPATH, "//option[. = '"+centroun+"']").click()
	driver.find_element(By.ID, "person-universityCenterId").click()
	carreraest=estu['carrera']
	driver.find_element(By.ID, "person-careerId").click()
	dropdown = driver.find_element(By.ID, "person-careerId")
	dropdown.find_element(By.XPATH, "//option[. = '"+carreraest+"']").click()
	driver.find_element(By.ID, "person-careerId").click()
	driver.find_element(By.CSS_SELECTOR, "main").click()
	driver.find_element(By.CSS_SELECTOR, ".text-right > .btn").click()
	#driver.implicitly_wait(50)
	##driver.execute_script("window.scrollTo(0,0)") # seconds
	#driver.find_element(By.ID, "perso-acceptTedff")
	#driver.find_element(By.ID, "person-acceptTerms").click()
	#driver.find_element(By.CSS_SELECTOR, ".btn:nth-child(2)").click()

def proc_insc(estu):
	agre_iden(estu)



""" *****  CUERPO PRINCIPAL ***** """
url_ambiente = 'https://srv-si-001.utpl.edu.ec/INSCRIPCION_NUEVOS_MAD/'


# Lectura de datos desde Excel
doc = openpyxl.load_workbook('plantilla_validacion.xlsx')
hoja = doc.get_sheet_by_name('Hoja1')
list_estu = recu_filas_excel(hoja, 1)

#Creacion de archivo de logs
f = open("ARCHIVO DE LOG "+obte_fech_actu()+".csv", "w+")
#f = open("ARCHIVO DE LOG.csv", "w+")
f.write("NUM FILA (EXCEL), CEDULA EST, OPERACION\n")

#Configuracion del navergador para pruebas
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(chrome_options=chrome_options)

class testadd(unittest.TestCase):
    def Ejecucionprueba(self):
        print("**** INICIO DEL PROCESAMIENTO DE INSCRIPCIONES ****")

        for estu in list_estu:
        	try:
        		print(
        			str(estu['nume_fila']) +" | ESTUDIANTE: " +
        			str(estu['TipoIden'])
        		)
        		login_site_test(driver)
        		proc_insc(estu)		
        	except Exception as e:
        		# escribir error
        		print("Error")
        		escribir_error_log(estu, "MANEGADOR PRICIPAL", str(e))
        	finally:
        		#cerrar el archivo de log
        		pass
        	
        f.close()

        print("**** FIN DEL PROCESO ****")
    def test_add1(self):
        self.assertEqual(2 + 3 + 10,15)
    def test_add2(self):
        self.assertEqual(10 + 150,160)
    def test_add3(self):
        #One error, view test results
        self.assertEqual(2 * 5 * 7, 40)
    def tearDown(self):
        pass
def suite():
    suiteTest=unittest.TestSuite()
    suiteTest.addTest(testadd("Ejecucionprueba"))
    #suiteTest.addTest(testadd("test_add2"))
    #suiteTest.addTest(testadd("test_add3"))
    #suiteTest.addTest("agre_iden")
    return suiteTest
if __name__=="__main__":
   # Store path in E Disk directory
   filepath='ResultadoEjecucion.html'
   fp=open(filepath,'wb')
   #Define the title and description of the test report
   runner = HTMLTestRunner.HTMLTestRunner(stream=fp,title=u'SOFTSERVICE',description=u'Resultado de la ejecución de la prueba de software')
   runner.run(suite())
   fp.close()